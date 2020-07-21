from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import json
import numpy as np
import xlrd
import os
from cv2 import cv2 as cv
from datetime import date,datetime
import time
from darkmodel import predictbox


#讀取數據資料
def data_loading(path_name):
    
    filename = os.listdir(r"./" + path_name)   
    label_data = []
    filename.sort(key=lambda x: int(x.split('.')[0]))
    array_list = []#最後輸出的陣列，共8400個70*11
    #最後輸出的標籤，共8400個
    date_string = ['20050103','20050331','20050401','20050630','20050701','20050930','20051003','20051230','20060102','20060331','20060403','20060630','20060703','20060929','20061002','20061229',
    '20070102','20070330','20070402','20070629','20070702','20070929','20071001','20071231','20080102','20080331','20080401','20080630','20080701','20080930','20081001','20081231','20090105',
    '20090331','20090401','20090630','20090701','20090930','20091001','20091231','20100104','20100331','20100401','20100630','20100701','20100930','20101001','20101231','20110103','20110331',
    '20110401','20110630','20110701','20110930','20111003','20111230','20120102','20120330','20120402','20120629','20120702','20120928','20121001','20121228','20130102','20130329','20130401',
    '20130628','20130701','20130930','20131001','20131231','20140102','20140331','20140401','20140630','20140701','20140930','20141001','20141231','20150105','20150331','20150401','20150630',
    '20150701','20150930','20151001','20151231','20160104','20160331','20160401','20160630','20160701','20160930','20161003','20161230','20170103','20170331','20170405','20170630','20170703',
    '20170930','20171002','20171229','20180102','20180331','20180402','20180629','20180702','20180928','20181001','20181228','20190102','20190329']
    #根據資料夾內的Excel檔名讀取
    count = 0
    for name in filename:
        
        data = xlrd.open_workbook(filename=r'./' + path_name +'/' + name)
        temp = 1
        row_temp = 0
        sheet1 = data.sheet_by_index(0)#讀檔案中第一個工作表
        array_list_temp = []#暫存每隻股票的陣列，時間從2005/1/3-2019/3/29
        label_temp = []#暫存每隻股票的標籤，時間從2005/1/3-2019/3/29


        print ("filename:", name)#顯示檔名
        print ("Work Sheet Rows:", sheet1.nrows)#顯示列數
        print ("Work Sheet Cols:", sheet1.ncols)#顯示行數
        
        label_first = 0#第一天之收盤價
        label_first = sheet1.cell(2,4).value
        label_last = 0#最後一天之收盤價
        array = np.zeros( (70,11) )#初始70*11的陣列為0

        #讀取檔案中每一行
        for i in range(2,sheet1.nrows):#讀取

            #轉換日期，原本為2005/1/3 00:00:00，轉換為20050103
            data_value=xlrd.xldate_as_tuple(sheet1.cell(i,0).value,data.datemode)
            data_value = (date(*data_value[:3])).strftime('%Y%m%d')

            #將檔案中最高價、最低價、開盤價、收盤價、均線(6種)以及KD線的資料讀入陣列中
            for x in range(1,5):
                array[row_temp][x-1] = sheet1.cell(i,x).value

            for y in range(7,14):
                array[row_temp][y-3] = sheet1.cell(i,y).value
                
            row_temp = row_temp + 1 
            
            #若讀到3個月的最後一天則執行
            if data_value == date_string[temp]:
                label_last = sheet1.cell(i,4).value#最後一天之收盤價
                row_temp = 0
                temp = temp + 2

                #計算上漲或下跌
                if label_last - label_first <= 0 :
                    label_temp.append(0)
                else:
                    label_temp.append(1)

                #將array加入array_list_temp中並初始為0
                if i != (sheet1.nrows-1):           
                    array_list_temp.append(array)
                    array = [[0]*11 for w in range(70)]
                    label_first = sheet1.cell(i+1,4).value

        
                    


        #因2005/1/3~2005/3/31的標籤不需要，將label_temp第一格刪除，2005/1/3~2005/3/31的標籤是由2005/6/30的收盤價-2005/4/1的收盤價得來的
        del label_temp[0]
        count = count -1
        array_list.extend(array_list_temp)#將每一支股票資料串接
        print(len(array_list))
        label_data.extend(label_temp)

    images,label,boxlist = load_dataset("test/") 


    return array_list,images,label,boxlist
    



def read_path(path_name):
    #讀取路經中的檔案，若資料夾有資料夾則遞迴讀取    
    dir_items = os.listdir(path_name)
    dir_items.sort(key=lambda x: (x.split('-')[1]))
    dir_items.sort(key=lambda x: (x.split('-')[0]))
    images = []

        
    for dir_item in dir_items:
        full_path = os.path.abspath(os.path.join(path_name, dir_item))        
        if os.path.isdir(full_path):   
            read_path(full_path)
        else:   
            if dir_item.endswith('.png'):
                #讀取圖片資料
                image = cv.imread(full_path)                 
                images.append(image)                                              
                     
    return images
def load_dataset(path_name):
    print(path_name)
    label = []
    images = read_path(path_name)#read_path return 所有圖片的list及圖片full path names
    
    boxlist =  predictbox(images) 
    images = np.array(images)
    
    #標註數據
    labels = os.listdir(path_name)
    labels.sort(key=lambda x: (x.split('-')[1]))
    labels.sort(key=lambda x: (x.split('-')[0]))
    for a in labels: 
        temp = os.path.splitext(a)[0]
        label.append(temp.split("-")[3])
    print(len(label))
    
    return images, label, boxlist
    
